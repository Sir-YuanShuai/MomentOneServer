from __future__ import annotations

import hashlib
import io
import json
from dataclasses import dataclass
from datetime import UTC, datetime
from decimal import Decimal, InvalidOperation
from typing import Any

from openpyxl import Workbook, load_workbook

YIMU_REQUIRED_HEADERS = {"日期", "收支类型", "金额", "类别", "账户", "账本"}
YIMU_EXPORT_HEADERS = (
    "日期",
    "收支类型",
    "金额",
    "类别",
    "二级分类",
    "账户",
    "账本",
    "退款",
    "优惠",
    "备注",
    "标签",
    "报销账户",
    "报销金额",
    "报销明细",
    "多币种",
    "地址",
    "创建用户",
    "其他",
    "附件1",
    "附件2",
    "附件3",
    "附件4",
    "附件5",
)
METADATA_HEADERS = (
    "Moment ID",
    "标题",
    "描述",
    "发生时间",
    "时区",
    "收支类型",
    "金额",
    "币种",
    "分类",
    "账户",
    "账本",
    "支付方式",
    "商家",
    "计入收支",
    "计入预算",
    "标签",
    "地点",
    "完整 Payload(JSON)",
    "创建时间",
    "更新时间",
    "Revision",
)


@dataclass(frozen=True, slots=True)
class ParsedRow:
    row_number: int
    title: str
    description: str | None
    occurred_at: datetime
    tags: tuple[str, ...]
    location_name: str | None
    payload: dict[str, Any]


@dataclass(frozen=True, slots=True)
class ParsedWorkbook:
    format_key: str
    rows: tuple[ParsedRow, ...]
    skipped_rows: int
    errors: tuple[dict[str, Any], ...]


def file_digest(content: bytes) -> str:
    return hashlib.sha256(content).hexdigest()


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _amount(value: Any) -> float | None:
    raw = _text(value).replace(",", "").replace("￥", "").replace("¥", "").replace("元", "")
    try:
        return abs(float(Decimal(raw)))
    except (InvalidOperation, ValueError):
        return None


def _date(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value.replace(tzinfo=value.tzinfo or UTC)
    raw = _text(value).replace("/", "-")
    if not raw:
        return None
    try:
        parsed = datetime.fromisoformat(raw.replace("Z", "+00:00"))
        return parsed.replace(tzinfo=parsed.tzinfo or UTC)
    except ValueError:
        return None


def _flow(value: Any) -> str | None:
    raw = _text(value).lower()
    if raw in {"支出", "expense", "支"}:
        return "expense"
    if raw in {"收入", "income", "收"}:
        return "income"
    return None


def parse_workbook(content: bytes) -> ParsedWorkbook:
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError("无法读取这个 Excel 文件。") from exc
    sheet = workbook["账单"] if "账单" in workbook.sheetnames else workbook.active
    if sheet is None:
        raise ValueError("Excel 中没有可读取的工作表。")
    rows = sheet.iter_rows(values_only=True)
    headers = tuple(_text(value) for value in next(rows, ()))
    if not YIMU_REQUIRED_HEADERS.issubset(set(headers)):
        raise ValueError("暂不识别这个 Excel 的表格结构。")
    indexes = {header: index for index, header in enumerate(headers) if header}
    parsed: list[ParsedRow] = []
    errors: list[dict[str, Any]] = []
    skipped = 0
    for row_number, values in enumerate(rows, start=2):

        def cell(name: str, current_values: tuple[Any, ...] = values) -> Any:
            index = indexes.get(name)
            return (
                current_values[index] if index is not None and index < len(current_values) else None
            )

        if not any(_text(value) for value in values):
            skipped += 1
            continue
        amount = _amount(cell("金额"))
        flow = _flow(cell("收支类型"))
        occurred_at = _date(cell("日期"))
        if amount is None or flow is None or occurred_at is None:
            errors.append({"rowNumber": row_number, "message": "日期、收支类型或金额无法识别"})
            continue
        primary = _text(cell("类别"))
        secondary = _text(cell("二级分类"))
        note = _text(cell("备注"))
        title = (secondary or primary or note or ("收入" if flow == "income" else "支出"))[:20]
        tags = tuple(filter(None, (_text(cell("标签")).replace("，", "、").split("、"))))[:5]
        payload = {
            "amount": amount,
            "flow": flow,
            "currency": _text(cell("多币种")) or "CNY",
            "category": secondary or primary or None,
            "account": _text(cell("账户")) or None,
            "ledger": _text(cell("账本")) or None,
            "merchant": _text(cell("地址")) or None,
            "countInFlow": True,
            "countInBudget": True,
            "importSource": "yimu",
            "sourceCategory": primary or None,
            "sourceSecondaryCategory": secondary or None,
            "refund": _text(cell("退款")) or None,
            "discount": _text(cell("优惠")) or None,
            "sourceCreator": _text(cell("创建用户")) or None,
            "sourceExtra": _text(cell("其他")) or None,
        }
        payload = {key: value for key, value in payload.items() if value is not None}
        parsed.append(
            ParsedRow(
                row_number,
                title,
                note[:240] or None,
                occurred_at,
                tags,
                _text(cell("地址")) or None,
                payload,
            )
        )
    return ParsedWorkbook("yimu", tuple(parsed), skipped, tuple(errors))


def export_workbook(moments: list[dict[str, Any]]) -> bytes:
    workbook = Workbook()
    bills = workbook.active
    if bills is None:
        raise RuntimeError("无法创建 Excel 工作表。")
    bills.title = "账单"
    bills.append(YIMU_EXPORT_HEADERS)
    metadata = workbook.create_sheet("Moment One 元数据")
    metadata.append(METADATA_HEADERS)
    for moment in moments:
        payload = moment.get("payload") or {}
        bills.append(
            (
                moment["occurredAt"],
                "收入" if payload.get("flow") == "income" else "支出",
                payload.get("amount"),
                payload.get("sourceCategory") or payload.get("category"),
                payload.get("sourceSecondaryCategory"),
                payload.get("account"),
                payload.get("ledger"),
                payload.get("refund"),
                payload.get("discount"),
                moment.get("description"),
                "、".join(moment.get("tags") or []),
                payload.get("reimbursementAccount"),
                payload.get("reimbursementAmount"),
                payload.get("reimbursementDetail"),
                payload.get("currency") or "CNY",
                (moment.get("location") or {}).get("name"),
                payload.get("sourceCreator"),
                payload.get("sourceExtra"),
            )
        )
        metadata.append(
            (
                moment["id"],
                moment["title"],
                moment.get("description"),
                moment["occurredAt"],
                moment["timezone"],
                payload.get("flow"),
                payload.get("amount"),
                payload.get("currency") or "CNY",
                payload.get("category"),
                payload.get("account"),
                payload.get("ledger"),
                payload.get("method"),
                payload.get("merchant"),
                payload.get("countInFlow", True),
                payload.get("countInBudget", True),
                "、".join(moment.get("tags") or []),
                (moment.get("location") or {}).get("name"),
                json.dumps(payload, ensure_ascii=False),
                moment["createdAt"],
                moment["updatedAt"],
                moment["revision"],
            )
        )
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()
