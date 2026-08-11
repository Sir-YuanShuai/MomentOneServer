from __future__ import annotations

import io
import json
from collections.abc import Iterable
from typing import Any

from openpyxl import Workbook

from app.modules.habit_goals.domain import HabitGoal
from app.modules.moments.domain import Moment

MOMENT_HEADERS = (
    "记录 ID",
    "标题",
    "描述",
    "发生时间",
    "时区",
    "分类",
    "标签",
    "地点",
    "场景数据(JSON)",
    "创建时间",
    "更新时间",
    "Revision",
)

HABIT_HEADERS = (
    "数据类型",
    "ID",
    "习惯/标题",
    "单位",
    "频率",
    "每周次数",
    "是否完成",
    "完成数量",
    "发生时间",
    "备注",
    "场景数据(JSON)",
    "创建时间",
    "更新时间",
    "Revision",
)


def _moment_row(moment: Moment) -> tuple[Any, ...]:
    return (
        str(moment.id),
        moment.title,
        moment.description,
        moment.occurred_at.isoformat(),
        moment.timezone,
        moment.category.value,
        "、".join(moment.tags),
        moment.location.name if moment.location else None,
        json.dumps(moment.payload, ensure_ascii=False),
        moment.created_at.isoformat(),
        moment.updated_at.isoformat(),
        moment.revision,
    )


def _append_moment_sheet(workbook: Workbook, title: str, moments: Iterable[Moment]) -> None:
    sheet = workbook.create_sheet(title)
    sheet.append(MOMENT_HEADERS)
    for moment in moments:
        sheet.append(_moment_row(moment))


def _append_habit_sheet(
    workbook: Workbook, goals: Iterable[HabitGoal], records: Iterable[Moment]
) -> None:
    sheet = workbook.create_sheet("习惯")
    sheet.append(HABIT_HEADERS)
    for goal in goals:
        sheet.append(
            (
                "目标",
                str(goal.id),
                goal.name,
                goal.unit,
                goal.frequency,
                goal.times_per_week,
                None,
                None,
                None,
                None,
                None,
                goal.created_at.isoformat(),
                goal.updated_at.isoformat(),
                goal.revision,
            )
        )
    for record in records:
        payload = record.payload
        sheet.append(
            (
                "完成记录",
                str(record.id),
                payload.get("habit") or record.title,
                payload.get("unit"),
                None,
                None,
                payload.get("done"),
                payload.get("count"),
                record.occurred_at.isoformat(),
                record.description,
                json.dumps(payload, ensure_ascii=False),
                record.created_at.isoformat(),
                record.updated_at.isoformat(),
                record.revision,
            )
        )


def export_personal_workbook(
    *, moments: Iterable[Moment], habit_goals: Iterable[HabitGoal], only: str | None = None
) -> bytes:
    """每个业务场景一个 Sheet；新增场景时只需在这里追加映射。"""
    all_moments = list(moments)
    workbook = Workbook()
    default = workbook.active
    if default is not None:
        workbook.remove(default)

    by_type: dict[str, list[Moment]] = {}
    for moment in all_moments:
        by_type.setdefault(moment.moment_type or "general", []).append(moment)

    if only == "habit":
        _append_habit_sheet(workbook, habit_goals, by_type.get("habit", []))
    else:
        _append_moment_sheet(workbook, "通用", by_type.get("general", []))
        _append_moment_sheet(workbook, "记账", by_type.get("bookkeeping", []))
        _append_habit_sheet(workbook, habit_goals, by_type.get("habit", []))
        for moment_type, records in sorted(by_type.items()):
            if moment_type not in {"general", "bookkeeping", "habit"}:
                _append_moment_sheet(workbook, moment_type[:31], records)

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()
