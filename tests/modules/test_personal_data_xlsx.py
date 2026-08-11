from datetime import UTC, datetime
from io import BytesIO
from uuid import uuid4

from app.modules.data_transfer.personal_data_xlsx import export_personal_workbook
from app.modules.habit_goals.domain import HabitGoal
from app.modules.moments.domain import Moment, MomentCategory
from openpyxl import load_workbook


def _moment(moment_type: str, title: str, payload: dict) -> Moment:
    now = datetime.now(UTC)
    return Moment(
        id=uuid4(),
        user_id=uuid4(),
        title=title,
        description="备注",
        voice_input=None,
        ai_summary=None,
        category=MomentCategory.EXPERIENCE,
        tags=("测试",),
        occurred_at=now,
        timezone="Asia/Shanghai",
        revision=1,
        created_at=now,
        updated_at=now,
        moment_type=moment_type,
        payload=payload,
    )


def test_all_data_export_uses_one_sheet_per_scene() -> None:
    now = datetime.now(UTC)
    goal = HabitGoal(
        id=uuid4(),
        user_id=uuid4(),
        name="阅读",
        unit="页",
        frequency="daily",
        times_per_week=7,
        color=None,
        revision=1,
        created_at=now,
        updated_at=now,
    )
    content = export_personal_workbook(
        moments=[
            _moment("general", "随手记", {}),
            _moment("bookkeeping", "午餐", {"amount": 25, "flow": "expense"}),
            _moment("habit", "阅读", {"habit": "阅读", "done": True, "count": 20}),
        ],
        habit_goals=[goal],
    )
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    assert workbook.sheetnames == ["通用", "记账", "习惯"]
    assert workbook["通用"].max_row == 2
    assert workbook["记账"].max_row == 2
    assert workbook["习惯"].max_row == 3


def test_habit_export_contains_goals_and_completion_records() -> None:
    now = datetime.now(UTC)
    goal = HabitGoal(
        id=uuid4(),
        user_id=uuid4(),
        name="喝水",
        revision=1,
        created_at=now,
        updated_at=now,
    )
    content = export_personal_workbook(
        moments=[_moment("habit", "喝水", {"habit": "喝水", "done": True})],
        habit_goals=[goal],
        only="habit",
    )
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    assert workbook.sheetnames == ["习惯"]
    assert [workbook["习惯"].cell(row=index, column=1).value for index in (2, 3)] == [
        "目标",
        "完成记录",
    ]
