import io

import pytest
from app.modules.data_transfer.bookkeeping_xlsx import export_workbook, parse_workbook
from openpyxl import Workbook, load_workbook


def workbook_bytes(headers: list[str], values: list[object] | None = None) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "账单"
    sheet.append(headers)
    if values is not None:
        sheet.append(values)
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def test_detects_yimu_by_sheet_structure_and_maps_row() -> None:
    headers = ["日期", "收支类型", "金额", "类别", "二级分类", "账户", "账本", "备注"]
    parsed = parse_workbook(
        workbook_bytes(
            headers, ["2026-08-10 12:00", "支出", 28.5, "餐饮", "午餐", "微信", "日常", "面馆"]
        )
    )
    assert parsed.format_key == "yimu"
    assert len(parsed.rows) == 1
    assert parsed.rows[0].title == "午餐"
    assert parsed.rows[0].payload["amount"] == 28.5
    assert parsed.rows[0].payload["flow"] == "expense"


def test_rejects_unknown_structure() -> None:
    with pytest.raises(ValueError, match="暂不识别"):
        parse_workbook(workbook_bytes(["时间", "钱", "说明"]))


def test_export_contains_compatible_and_complete_sheets() -> None:
    content = export_workbook(
        [
            {
                "id": "moment-1",
                "title": "午餐",
                "description": "面馆",
                "occurredAt": "2026-08-10T12:00:00+08:00",
                "timezone": "Asia/Shanghai",
                "payload": {"amount": 28.5, "flow": "expense", "currency": "CNY"},
                "tags": ["工作日"],
                "location": None,
                "createdAt": "2026-08-10T12:00:00+08:00",
                "updatedAt": "2026-08-10T12:00:00+08:00",
                "revision": 1,
            }
        ]
    )
    workbook = load_workbook(io.BytesIO(content), read_only=True)
    assert workbook.sheetnames == ["账单", "Moment One 元数据"]
    assert workbook["账单"].max_row == 2
    assert workbook["Moment One 元数据"].max_row == 2
